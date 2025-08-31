"""
CSV and Excel export functionality for leads
"""
import csv
import pandas as pd
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime
from loguru import logger

from src.models.lead import Lead
from config.settings import EXPORTS_DIR


class CSVExporter:
    """Handles CSV and Excel export of leads"""
    
    def __init__(self, output_dir: Path = EXPORTS_DIR):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
    
    def export_to_csv(self, leads: List[Lead], filename: str = None) -> str:
        """Export leads to CSV file"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"leads_export_{timestamp}.csv"
        
        filepath = self.output_dir / filename
        
        try:
            # Convert leads to dictionaries
            lead_data = [lead.to_dict() for lead in leads]
            
            if not lead_data:
                logger.warning("No leads to export")
                return str(filepath)
            
            # Get all possible fieldnames
            fieldnames = set()
            for lead_dict in lead_data:
                fieldnames.update(lead_dict.keys())
            
            fieldnames = sorted(list(fieldnames))
            
            # Write CSV
            with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(lead_data)
            
            logger.info(f"Exported {len(leads)} leads to CSV: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error exporting to CSV: {e}")
            raise
    
    def export_to_excel(self, leads: List[Lead], filename: str = None, include_analytics: bool = True) -> str:
        """Export leads to Excel file with multiple sheets"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"leads_export_{timestamp}.xlsx"
        
        filepath = self.output_dir / filename
        
        try:
            # Convert leads to DataFrame
            lead_data = [lead.to_dict() for lead in leads]
            df_leads = pd.DataFrame(lead_data)
            
            if df_leads.empty:
                logger.warning("No leads to export")
                return str(filepath)
            
            # Create Excel writer
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Main leads sheet
                df_leads.to_excel(writer, sheet_name='Leads', index=False)
                
                if include_analytics:
                    # Analytics sheets
                    self._add_analytics_sheets(writer, df_leads)
                
                # Format sheets
                self._format_excel_sheets(writer, df_leads)
            
            logger.info(f"Exported {len(leads)} leads to Excel: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            raise
    
    def export_by_platform(self, leads: List[Lead], base_filename: str = None) -> Dict[str, str]:
        """Export leads grouped by platform to separate files"""
        if not base_filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_filename = f"leads_by_platform_{timestamp}"
        
        # Group leads by platform
        platform_leads = {}
        for lead in leads:
            platform = lead.platform
            if platform not in platform_leads:
                platform_leads[platform] = []
            platform_leads[platform].append(lead)
        
        exported_files = {}
        
        # Export each platform separately
        for platform, platform_lead_list in platform_leads.items():
            filename = f"{base_filename}_{platform}.xlsx"
            filepath = self.export_to_excel(platform_lead_list, filename)
            exported_files[platform] = filepath
        
        logger.info(f"Exported leads to {len(exported_files)} platform-specific files")
        return exported_files
    
    def export_high_score_leads(self, leads: List[Lead], min_score: int = 50, filename: str = None) -> str:
        """Export only high-scoring leads"""
        high_score_leads = [lead for lead in leads if lead.lead_score >= min_score]
        
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"high_score_leads_{min_score}+_{timestamp}.xlsx"
        
        logger.info(f"Exporting {len(high_score_leads)} leads with score >= {min_score}")
        return self.export_to_excel(high_score_leads, filename)
    
    def _add_analytics_sheets(self, writer: pd.ExcelWriter, df_leads: pd.DataFrame):
        """Add analytics sheets to Excel export"""
        try:
            if 'platform' in df_leads.columns:
                # Platform distribution
                platform_stats = df_leads['platform'].value_counts().reset_index()
                platform_stats.columns = ['Platform', 'Count']
                platform_stats.to_excel(writer, sheet_name='Platform Stats', index=False)
            else:
                pd.DataFrame({'Platform': [], 'Count': []}).to_excel(writer, sheet_name='Platform Stats', index=False)
            
            # Lead score distribution
            if 'lead_score' in df_leads.columns:
                score_ranges = pd.cut(df_leads['lead_score'], bins=[0, 20, 40, 60, 80, 100], 
                                    labels=['0-20', '21-40', '41-60', '61-80', '81-100'])
                score_stats = score_ranges.value_counts().reset_index()
                score_stats.columns = ['Score Range', 'Count']
                score_stats.to_excel(writer, sheet_name='Score Distribution', index=False)
            else:
                pd.DataFrame({'Score Range': [], 'Count': []}).to_excel(writer, sheet_name='Score Distribution', index=False)
            
            # Industry distribution (if available)
            if 'industry' in df_leads.columns and not df_leads['industry'].isna().all():
                industry_stats = df_leads['industry'].value_counts().head(10).reset_index()
                industry_stats.columns = ['Industry', 'Count']
                industry_stats.to_excel(writer, sheet_name='Top Industries', index=False)
            else:
                pd.DataFrame({'Industry': [], 'Count': []}).to_excel(writer, sheet_name='Top Industries', index=False)
            
            # Contact information availability
            contact_stats = {
                'Has Email': int(df_leads['email'].notna().sum()) if 'email' in df_leads.columns else 0,
                'Has Phone': int(df_leads['phone'].notna().sum()) if 'phone' in df_leads.columns else 0,
                'Has Website': int(df_leads['website'].notna().sum()) if 'website' in df_leads.columns else 0,
                'Has Address': int(df_leads['address'].notna().sum()) if 'address' in df_leads.columns else 0,
                'Has Social Handles': int(df_leads['social_handles'].notna().sum()) if 'social_handles' in df_leads.columns else 0
            }
            
            contact_df = pd.DataFrame(list(contact_stats.items()), 
                                    columns=['Contact Type', 'Count'])
            contact_df.to_excel(writer, sheet_name='Contact Info Stats', index=False)
            
            # Top pain points
            all_pain_points = []
            if 'pain_points' in df_leads.columns:
                for pain_points_str in df_leads['pain_points'].dropna():
                    if pain_points_str:
                        all_pain_points.extend(pain_points_str.split(', '))
            
            if all_pain_points:
                pain_point_stats = pd.Series(all_pain_points).value_counts().head(10).reset_index()
                pain_point_stats.columns = ['Pain Point', 'Count']
                pain_point_stats.to_excel(writer, sheet_name='Top Pain Points', index=False)
            
        except Exception as e:
            logger.warning(f"Error creating analytics sheets: {e}")
    
    def _format_excel_sheets(self, writer: pd.ExcelWriter, df_leads: pd.DataFrame):
        """Format Excel sheets for better readability"""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            workbook = writer.book
            
            # Format main leads sheet
            if 'Leads' in workbook.sheetnames:
                worksheet = workbook['Leads']
                
                # Header formatting
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Freeze header row
                worksheet.freeze_panes = "A2"
            
        except Exception as e:
            logger.warning(f"Error formatting Excel sheets: {e}")
    
    def create_lead_summary_report(self, leads: List[Lead], filename: str = None) -> str:
        """Create a comprehensive summary report"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"lead_summary_report_{timestamp}.xlsx"
        
        filepath = self.output_dir / filename
        
        try:
            df_leads = pd.DataFrame([lead.to_dict() for lead in leads])
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Executive Summary
                summary_metrics = [
                    ('Total Leads', len(leads)),
                    ('Average Lead Score', f"{df_leads['lead_score'].mean():.1f}" if 'lead_score' in df_leads.columns and not df_leads.empty else 'N/A'),
                    ('High Score Leads (60+)', len([l for l in leads if l.lead_score >= 60])),
                    ('Leads with Email', int(df_leads['email'].notna().sum()) if 'email' in df_leads.columns else 0),
                    ('Leads with Phone', int(df_leads['phone'].notna().sum()) if 'phone' in df_leads.columns else 0),
                    ('Leads with Website', int(df_leads['website'].notna().sum()) if 'website' in df_leads.columns else 0),
                    ('Most Common Platform', df_leads['platform'].mode().iloc[0] if 'platform' in df_leads.columns and not df_leads.empty else 'N/A'),
                    ('Most Common Industry', df_leads['industry'].mode().iloc[0] if 'industry' in df_leads.columns and not df_leads['industry'].isna().all() else 'N/A')
                ]
                summary_df = pd.DataFrame(summary_metrics, columns=['Metric', 'Value'])
                summary_df.to_excel(writer, sheet_name='Executive Summary', index=False)
                # Full leads data
                df_leads.to_excel(writer, sheet_name='All Leads', index=False)
                # Add analytics
                self._add_analytics_sheets(writer, df_leads)
                # Format
                self._format_excel_sheets(writer, df_leads)
            
            logger.info(f"Created summary report: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error creating summary report: {e}")
            raise
